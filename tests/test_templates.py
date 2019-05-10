import os
import io
import decimal
import itertools
import string

from openpyxl import load_workbook

from xlsx_template.template import Template
import data_generators


BASE_TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "xlsx")


def test_variables(render_template):
    wb = render_template(
        "test_variables.xlsx", data_generators.generate_for_test_variables()
    )
    assert "Sheet My Sheet" in wb
    assert "10" in wb
    assert "Another sheet" in wb
    ws = wb["Sheet My Sheet"]
    assert ws["A1"].value == "String Constant"
    assert ws["A2"].value is True
    assert ws["A3"].value is False
    assert ws["A4"].value == 123
    assert ws["A5"].value == 123.45
    assert ws["A6"].value == "String"
    assert ws["A7"].value == 10
    assert ws["A8"].value == 10.0
    assert ws["A9"].value == 10.1
    assert ws["A10"].value == "Complex var 10"
    assert ws["A11"].value is None
    assert ws["A12"].value == "simple_call"
    assert ws["A13"].value == "1, True, 123"
    assert ws["A14"].value == "kwarg1=1, kwarg2=123, kwarg3=False"
    assert ws["A15"].value == "s, kwarg1=True"
    assert ws["A16"].value is None
    assert ws["A17"].value == "getattr attr1"
    assert ws["A18"].value == "getitem 'some_key'"
    assert ws["A19"].value == "Variable is None"


def test_simple_loop(render_template):
    wb = render_template(
        "test_simple_loop.xlsx", data_generators.generate_for_test_simple_loop()
    )
    ws = wb["Sheet1"]
    assert ws["A1"].value == "Top"
    for row in range(2, 7):
        assert ws["A{}".format(row)].value == "item{}".format(row - 1)
    assert ws["A7"].value == "Bottom"

    assert ws["A9"].value == "Left"
    for index, col in enumerate("BCDEF", 1):
        assert ws["{}9".format(col)].value == "item{}".format(index)


def _test_two_loops(wb):
    ws = wb["Sheet1"]
    rows = [str(_) for _ in range(2, 12)]
    cols = string.ascii_uppercase[1:11]
    for index, (row, col) in enumerate(itertools.product(rows, cols)):
        cell = col + row
        assert ws[cell].value == index
    for cell in ("A1", "L1", "A12", "L12"):
        assert ws[cell].value == "=SUM(B2:K11)"

    for row in rows:
        cell1 = "A" + row
        cell2 = "L" + row
        valid_value = "=SUM(B{0}:K{0})".format(row)
        assert ws[cell1].value == valid_value
        assert ws[cell2].value == valid_value

    for col in cols:
        cell1 = col + "1"
        cell2 = col + "12"
        valid_value = "=SUM({0}2:{0}11)".format(col)
        assert ws[cell1].value == valid_value
        assert ws[cell2].value == valid_value


def test_two_loops(render_template):
    _test_two_loops(
        render_template(
            "test_two_nested_loops.xlsx",
            data_generators.generate_for_test_two_nested_loops(),
        )
    )


def test_two_loops_synt_v2(render_template):
    _test_two_loops(
        render_template(
            "test_two_nested_loops_synt_v2.xlsx",
            data_generators.generate_for_test_two_nested_loops(),
        )
    )


def test_loop_context(render_template):
    data = data_generators.generate_for_test_loop_context()
    items_length = len(data["items"])
    wb = render_template("test_loop_context.xlsx", data)
    ws = wb["Sheet"]
    for index, row in enumerate(range(2, 7)):
        assert ws.cell(row, 1).value == ws.cell(row, 8).value == index + 1
        assert ws.cell(row, 2).value == ws.cell(row, 9).value == index
        assert ws.cell(row, 3).value == ws.cell(row, 10).value == items_length - index
        assert (
            ws.cell(row, 4).value == ws.cell(row, 11).value == items_length - index - 1
        )
        assert ws.cell(row, 5).value == ws.cell(row, 12).value == (index == 0)
        assert (
            ws.cell(row, 6).value
            == ws.cell(row, 13).value
            == (index == items_length - 1)
        )
        assert ws.cell(row, 7).value == ws.cell(row, 14).value == items_length
        assert ws.cell(row, 15).value == "item{}".format(index)


def test_merge(render_template):
    data = data_generators.generate_for_test_merge()
    wb = render_template("test_merge.xlsx", data)
    ws = wb["Sheet"]
    assert "A1:C2" in ws.merged_cells
    for row in range(5, 10):
        assert "A{0}:C{0}".format(row) in ws.merged_cells
    assert "B11:B13" in ws.merged_cells
    assert "A16:C16" in ws.merged_cells
    assert "A18:C20" in ws.merged_cells


def test_if(render_template):
    data = data_generators.generate_for_test_if()
    wb = render_template("test_if.xlsx", data)
    ws = wb["Sheet"]
    for row in range(3, 13):
        value1 = ws["A{}".format(row)].value
        value2 = "red" if value1 % 2 == 0 else "green"
        assert ws["B{}".format(row)].value == value2
