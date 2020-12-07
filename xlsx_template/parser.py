import copy
import io

import pyparsing
from openpyxl import load_workbook, styles

from . import grammar, nodes, utils
from .consts import LoopDirection, FuncArgDirection
from .exceptions import ParseError


class Parser:
    def __init__(self, source, filename=None):
        self.wb = load_workbook(io.BytesIO(source))
        self.source_hint = ["filename:{}".format(filename)] if filename else []

    def parse(self):
        template = nodes.Template(body=[])
        self.styles = {}
        for sheet_name in self.wb.sheetnames:
            self.source_hint.append("sheet:{}".format(sheet_name))
            ws = self.wb[sheet_name]
            self.cells = {}
            self.merged_cells = {}
            self.post_remove = []
            self.original_cell_groups = {}
            for cr in ws.merged_cells:
                self.merged_cells[(cr.min_row, cr.min_col)] = (
                    cr.max_row - cr.min_row + 1,
                    cr.max_col - cr.min_col + 1,
                )
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    self.source_hint.append(
                        "cell:{}".format(utils.cell_int_to_str(row, col))
                    )
                    node = self.parse_cell(ws, row, col)
                    self.source_hint.pop()
                    self.cells[(row, col)] = node
            directives = self.get_directives(ws)
            self.directives = self.parse_directives(directives)
            body = self._process_cell_group(1, 1, ws.max_row, ws.max_column)
            self.source_hint.append("sheet name")
            sheet_name_node = self.parse_value(sheet_name)
            self.source_hint.pop()
            if isinstance(body[0], nodes.SheetLoop):
                root_node = body[0]
                root_node.sheet.name=nodes.ToStr(value=sheet_name_node)
            else:
                root_node = nodes.Sheet(
                    name=nodes.ToStr(value=sheet_name_node),
                    max_row=ws.max_row,
                    max_col=ws.max_column,
                    body=body,
                )
            self.source_hint.pop()

            assert not self.source_hint
            if self.post_remove:
                for start_cell, end_cell in self.post_remove:
                    target_cell_group = None
                    target_cell = None
                    for (row, col), cell_group in self.original_cell_groups.items():
                        if (
                            start_cell[0] >= row
                            and start_cell[1] >= col
                            and end_cell[0] <= row + cell_group.height
                            and end_cell[1] <= col + cell_group.width
                        ):
                            target_cell_group = cell_group
                            target_cell = (row, col)
                    assert target_cell_group is not None
                    target_cell_group.body.append(
                        nodes.Remove(
                            base_cell=(
                                start_cell[0] - target_cell[0],
                                start_cell[1] - target_cell[1],
                            ),
                            last_cell=(
                                end_cell[0] - target_cell[0],
                                end_cell[1] - target_cell[1],
                            ),
                        )
                    )
            template.body.append(root_node)
            del self.post_remove
            del self.original_cell_groups
        return template, list(self.styles.values())

    def _process_cell_group(self, start_row, start_col, end_row, end_col):
        body = []
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                if (row, col) in self.directives and self.directives[(row, col)]:
                    cur_directives = self.directives[(row, col)]
                    cur_directive = cur_directives.pop(0)
                    if isinstance(cur_directive, nodes.SheetLoop):
                        cur_directive.last_cell = (end_row, end_col)
                    if not cur_directives:
                        del self.directives[(row, col)]
                    method = getattr(
                        self,
                        "_process_{}".format(cur_directive.__class__.__name__.lower()),
                    )
                    cur_directive = method(cur_directive)
                    if cur_directive is not None:
                        if isinstance(cur_directive, nodes.CellGroup):
                            self.original_cell_groups[(row, col)] = cur_directive
                        cur_directive.adjust(-start_row, -start_col)
                        body.append(cur_directive)
                if (row, col) in self.cells:
                    cell = self.cells.pop((row, col))
                    cell.adjust(-start_row, -start_col)
                    body.append(cell)
        return body

    def _process_cellloop(self, cell_loop):
        cell_loop.body = self._process_cell_group(
            cell_loop.base_cell[0],
            cell_loop.base_cell[1],
            cell_loop.last_cell[0],
            cell_loop.last_cell[1],
        )
        return cell_loop

    def parse_directives(self, directives):
        res_directives = {}
        for (row, col), directives in directives.items():
            self.source_hint.append("cell:{}".format(utils.cell_int_to_str(row, col)))
            res_directives[(row, col)] = []
            for index, directive in enumerate(directives):
                self.source_hint.append("directive_index:{}".format(index))
                res_directives[(row, col)].append(
                    self.parse_directive(row, col, directive)
                )
                self.source_hint.pop()
            self.source_hint.pop()
        return res_directives

    def parse_directive(self, row, col, directive):
        if "," in directive:
            directive_name, tail = directive.split(",", 1)
        else:
            directive_name, tail = directive, ""
        directive_name = directive_name.lower().replace("-", "_")
        method = getattr(self, "parse_{}".format(directive_name), None)
        if method is None:
            raise ParseError(
                "Invalid directive {}".format(directive_name), self.source_hint
            )
        node = method(tail)
        node.base_cell = (row, col)
        if hasattr(node, "last_cell"):
            if node.last_cell is None:
                node.last_cell = node.base_cell
            if len(node.last_cell) == 1:
                node.last_cell = node.last_cell[0]
        return node

    def parse_merge(self, directive_def):
        return self._parse_pp(grammar.parse_merge, directive_def)

    def _process_merge(self, merge):
        cell_output = self.cells[merge.base_cell]
        cell_output.merge = merge

    def parse_group(self, directive_def):
        return self._parse_pp(grammar.parse_group, directive_def)

    def parse_remove(self, directive_def):
        return self._parse_pp(grammar.parse_remove, directive_def)

    def _process_remove(self, remove):
        for cell in remove.get_cell_range():
            self.directives.pop(cell, None)
            self.cells.pop(cell, None)
        return remove

    def parse_if(self, directive_def):
        return self._parse_pp(grammar.parse_if, directive_def)

    def parse_col_width(self, directive_def):
        return self._parse_pp(grammar.parse_col_width, directive_def)

    def parse_row_height(self, directive_def):
        return self._parse_pp(grammar.parse_row_height, directive_def)

    def _process_colwidth(self, col_width):
        self.cells[col_width.base_cell].col_width = col_width.value

    def _process_rowheight(self, row_height):
        self.cells[row_height.base_cell].row_height = row_height.value

    def _process_if(self, if_d):
        if_d.body = self._process_cell_group(
            if_d.base_cell[0], if_d.base_cell[1], if_d.last_cell[0], if_d.last_cell[1]
        )
        if if_d.else_block:
            start_cell = if_d.else_block[0]
            if len(if_d.else_block) == 2:
                end_cell = if_d.else_block[1]
            else:
                end_cell = start_cell
            if_d.else_block = self._process_cell_group(
                start_cell[0], start_cell[1], end_cell[0], end_cell[1]
            )
            self.post_remove.append((start_cell, end_cell))
        return if_d

    def _process_cellgroup(self, cell_group):
        cell_group.body = self._process_cell_group(
            cell_group.base_cell[0],
            cell_group.base_cell[1],
            cell_group.last_cell[0],
            cell_group.last_cell[1],
        )
        return cell_group

    def _process_sheetloop(self, sheet_loop):
        sheet = nodes.Sheet(
            None,
            max_row=sheet_loop.last_cell[0],
            max_col=sheet_loop.last_cell[1],
            body=self._process_cell_group(
                sheet_loop.base_cell[0],
                sheet_loop.base_cell[1],
                sheet_loop.last_cell[0],
                sheet_loop.last_cell[1],
            ),
        )
        sheet_loop.sheet = sheet
        return sheet_loop

    def parse_func_arg_v(self, directive_def):
        return nodes.FuncArgDirection(direction=FuncArgDirection.VERTICAL)

    def parse_func_arg_h(self, directive_def):
        return nodes.FuncArgDirection(direction=FuncArgDirection.HORIZONTAL)

    def _process_funcargdirection(self, func_arg_dir):
        cell = (func_arg_dir.base_cell[0], func_arg_dir.base_cell[1])
        func_cell = self.cells.get(cell)
        assert isinstance(func_cell, nodes.FuncCellOutput), "Cell must contain function"
        for arg in func_cell.args:
            arg.direction = func_arg_dir.direction
        assert (
            self.directives.get(cell) is None
        ), "Func-arg-direction directive must be the last"

    def parse_loop_right(self, directive_def):
        return self._parse_cell_loop(directive_def, LoopDirection.RIGHT)

    def parse_loop_down(self, directive_def):
        return self._parse_cell_loop(directive_def, LoopDirection.DOWN)

    def _parse_cell_loop(self, directive_def, loop_direction):
        node = self._parse_pp(grammar.parse_cell_loop_directive, directive_def)
        node.direction = loop_direction
        if node.last_cell is None:
            node.last_cell = node.base_cell
        return node

    def parse_loop_sheet(self, directive_def):
        node = self._parse_pp(grammar.parse_sheet_loop_directive, directive_def)
        return node

    def get_directives(self, ws):
        directives = {}
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if cell.comment and cell.comment.text:
                    lines = [line.strip() for line in cell.comment.text.splitlines()]
                    lines = [line for line in lines if line]
                    if lines[0].lower() == "synt-v2":
                        v2_directives = self.get_directives_synt_v2(lines[1:])
                        for (row, col), cell_directives in v2_directives.items():
                            if (row, col) not in directives:
                                directives[(row, col)] = []
                            directives[(row, col)].extend(cell_directives)
                    else:
                        directives[(row, col)] = lines
        return directives

    def get_directives_synt_v2(self, lines):
        state = 0
        current_cell = None
        res_directives = {}
        for line in lines:
            if state == 0:
                current_cell = grammar.cell.parseString(line, True)[0]
                state = 1
            else:
                if all(c == "=" for c in line):
                    state = 0
                else:
                    if current_cell not in res_directives:
                        res_directives[current_cell] = []
                    res_directives[current_cell].append(line)
        return res_directives

    def _parse_pp(self, pp, s):
        try:
            return pp(s)
        except pyparsing.ParseBaseException as e:
            msg = "Error parse '{}'".format(s)
            raise ParseError(msg, self.source_hint, e)

    def parse_expression(self, expr):
        node = self._parse_pp(grammar.parse_expr_with_filter, expr)
        return node

    def parse_value(self, value):
        if not isinstance(value, str) or "{{" not in value:
            return nodes.Const(value=value)
        else:
            body = []
            index = 0
            while index < len(value):
                start_var_index = value.find("{{", index)
                # If value start with {{
                if start_var_index - index == 0:
                    end_var_index = value.find("}}", start_var_index)
                    if end_var_index != -1:
                        self.source_hint.append(
                            "index:{}-{}".format(start_var_index + 2, end_var_index)
                        )
                        body.append(
                            self.parse_expression(
                                value[start_var_index + 2 : end_var_index]
                            )
                        )
                        self.source_hint.pop()
                        index = end_var_index + 2
                    else:
                        body.append(nodes.Const(value=value[index:]))
                        index = len(value)
                else:
                    body.append(nodes.Const(value=value[:start_var_index]))
                    index = start_var_index
            if len(body) != 1:
                body = [
                    nodes.ToStr(value=child)
                    if not isinstance(child, nodes.Const)
                    else child
                    for child in body
                ]
            return nodes.Value(body=body)

    def parse_cell(self, ws, row, col):
        cell = ws.cell(row, col)
        style_name = None
        if cell.has_style:
            style_key = (
                copy.copy(cell.font),
                copy.copy(cell.fill),
                copy.copy(cell.border),
                copy.copy(cell.alignment),
                copy.copy(cell.number_format),
                copy.copy(cell.protection),
            )
            if style_key not in self.styles:
                new_style_name = "style{}".format(len(self.styles))
                new_style = styles.NamedStyle(new_style_name)
                new_style.font = style_key[0]
                new_style.fill = style_key[1]
                new_style.border = style_key[2]
                new_style.alignment = style_key[3]
                new_style.number_format = style_key[4]
                new_style.protection = style_key[5]
                self.styles[style_key] = new_style
                style_name = new_style_name
            else:
                style_name = self.styles[style_key].name
        col_letter = utils.col_int_to_str(col)
        col_width = ws.column_dimensions[col_letter].width
        col_index = col
        while not col_width:
            col_letter = utils.col_int_to_str(col_index)
            col_width = ws.column_dimensions[col_letter].width
            col_index -= 1
        node_kwargs = {
            "base_cell": (row, col),
            "style": style_name,
            "row_height": nodes.Const(value=ws.row_dimensions[row].height),
            "col_width": nodes.Const(value=col_width),
        }
        node_class = nodes.CellOutput
        if cell.value:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                value = cell.value
                node_kwargs["args"] = self.parse_func_args(value)
                node_class = nodes.FuncCellOutput
            else:
                value = self.parse_value(cell.value)
        else:
            value = None
        node_kwargs["value"] = value
        if (row, col) in self.merged_cells:
            node_kwargs["merge"] = nodes.Merge(
                rows=nodes.Const(value=self.merged_cells[(row, col)][0]),
                cols=nodes.Const(value=self.merged_cells[(row, col)][1]),
            )
        cell_output = node_class(**node_kwargs)
        return cell_output

    def parse_func_args(self, s):
        res = []
        for (start_index, end_index), cell_def in grammar.parse_func_args(s):
            if len(cell_def) == 1:
                cells = cell_def
            else:
                cells = [
                    (row, col)
                    for row in range(cell_def[0][0], cell_def[1][0] + 1)
                    for col in range(cell_def[0][1], cell_def[1][1] + 1)
                ]
            res.append(
                nodes.FuncArg(start_index=start_index, end_index=end_index, cells=cells)
            )
        return res
